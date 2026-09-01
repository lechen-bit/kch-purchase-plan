import json
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
INPUT = Path(os.environ.get("INPUT_XLSX", r"C:\Users\KCH\Downloads\Telegram Desktop\2026구매계획0513.xlsx"))
REFERENCE = ROOT / "reference-data.json"
STORE = ROOT / "data-store.json"
DATA_JS = ROOT / "data.js"

MONTHS = {
    "1": 1,
    "1월": 1,
    "jan": 1,
    "january": 1,
    "2": 2,
    "2월": 2,
    "feb": 2,
    "february": 2,
    "3": 3,
    "3월": 3,
    "mar": 3,
    "march": 3,
    "4": 4,
    "4월": 4,
    "apr": 4,
    "april": 4,
    "5": 5,
    "5월": 5,
    "may": 5,
    "6": 6,
    "6월": 6,
    "jun": 6,
    "june": 6,
    "7": 7,
    "7월": 7,
    "jul": 7,
    "july": 7,
    "8": 8,
    "8월": 8,
    "aug": 8,
    "august": 8,
    "9": 9,
    "9월": 9,
    "sep": 9,
    "september": 9,
    "10": 10,
    "10월": 10,
    "oct": 10,
    "october": 10,
    "11": 11,
    "11월": 11,
    "nov": 11,
    "november": 11,
    "12": 12,
    "12월": 12,
    "dec": 12,
    "december": 12,
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def month_number(value):
    text = clean(value).lower()
    return MONTHS.get(text)


def period_year_month(value, default_year=2026):
    text = clean(value).lower().replace(" ", "")
    match = re.match(r"(?:(\d{2,4})년)?(\d{1,2})월", text)
    if match:
        year_text, month_text = match.groups()
        year = default_year
        if year_text:
            year_num = int(year_text)
            year = 2000 + year_num if year_num < 100 else year_num
        month = int(month_text)
        if 1 <= month <= 12:
            return year, month
    month = month_number(value)
    return (default_year, month) if month else (None, None)


def month_dates(year, month):
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year, 12, 31)
    else:
        end = datetime(year, month + 1, 1)
        end = datetime.fromtimestamp(end.timestamp() - 86400)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def parse_laycan_note(year, note):
    text = clean(note)
    match = re.search(r"\b(\d{1,2})[./](\d{1,2})\s*[-~–]\s*(?:(\d{1,2})[./])?(\d{1,2})\b", text)
    if not match:
        return None
    start_month = int(match.group(1))
    start_day = int(match.group(2))
    end_month = int(match.group(3) or start_month)
    end_day = int(match.group(4))
    end_year = year + 1 if end_month < start_month else year
    try:
        start = datetime(year, start_month, start_day)
        end = datetime(end_year, end_month, end_day)
    except ValueError:
        return None
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def normalize_type(value):
    text = clean(value).lower()
    if text.startswith("fix") or text in {"first", "match"}:
        return "fixed"
    if text.startswith("index"):
        return "index"
    return "fixed"


def status_and_buyer(sell, end_user):
    sell_text = clean(sell)
    end_user_text = clean(end_user)
    if not sell_text or sell_text.lower() in {"unsold", "un sold", "for sale"}:
        return "forsale", ""
    if sell_text.lower() in {"planned", "plan"}:
        return "planned", ""
    buyer = end_user_text or re.sub(r"\d+\s*차$", "", sell_text).strip()
    return "sold", buyer


def parse_price(value, purchase_type):
    text = clean(value).replace("(CFR)", "").replace("CFR", "").strip()
    if purchase_type == "index":
        return None, text
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return (float(match.group()) if match else None), None


def normalize_rule(value):
    text = clean(value).lower()
    if "3" in text and "week" in text:
        return "3 weeks before laycan"
    if "4" in text and "week" in text:
        return "4 weeks before laycan"
    return "1 month before laycan"


def import_2026_cargos():
    wb = openpyxl.load_workbook(INPUT, data_only=True)
    # Priority 1: sheets whose name contains both "2026" and "요약" (e.g. "2026구매계획 요약")
    # Priority 2: sheets named exactly "구매계획" but ONLY if they look like the right format
    #             (row 2 col A == "Month" or row 1 col A == 2026/year number)
    # Priority 3: active sheet
    ws = None
    # try "2026…요약" style first
    for name in wb.sheetnames:
        if "2026" in name and "요약" in name:
            ws = wb[name]
            break
    # try exact template sheet name (our new template) — verify it has right format
    if ws is None and "구매계획" in wb.sheetnames:
        candidate = wb["구매계획"]
        # our template sheet starts data at row 5 (rows 1-4 are title/notice/header/hint)
        # and row 3 header col A == "기간"; the legacy "구매계획" sheet has col B=="Purchase"
        # Check: if row 2, col A is '기간' → our template; col A is 'Year' → wrong sheet
        r2a = str(candidate.cell(2, 1).value or "").strip().lower()
        r2b = str(candidate.cell(2, 2).value or "").strip().lower()
        if r2a == "기간" or (r2a not in ("year", "") and r2b in ("purchase", "매입유형", "fix", "index")):
            ws = candidate
    if ws is None:
        ws = wb.active

    # ICI sheet import — done here so we open the file only once
    _ici_from_wb = import_ici_from_workbook(wb)
    import_2026_cargos._ici_entries = _ici_from_wb

    cargos = []

    # Auto-detect first data row: scan rows 1-10, skip rows where col A looks like a header/year
    first_data_row = 3
    for probe_r in range(1, 11):
        val_a = ws.cell(probe_r, 1).value
        val_c = ws.cell(probe_r, 3).value
        if val_a is None:
            continue
        a_str = str(val_a).strip().lower()
        # Skip pure-year rows (e.g. 2026), header label rows, notice rows
        if a_str in ("", "month", "year", "2024", "2025", "2026", "2026년"):
            continue
        if a_str.startswith("※") or a_str.startswith("note"):
            continue
        # Row looks like data if col A is a month label and col C looks like a mine name
        if val_c and a_str not in ("purchase", "mine", "spec", "cost price"):
            first_data_row = probe_r
            break

    for row_index, row in enumerate(ws.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        year, month = period_year_month(row[0])
        purchase = clean(row[1])
        mine = clean(row[2])
        spec = clean(row[3])
        if not year or not month or not mine or not spec:
            continue

        purchase_type = normalize_type(purchase)
        price_fixed, price_formula = parse_price(row[4], purchase_type)
        quantity = int(float(clean(row[5]).replace(",", ""))) if clean(row[5]) else 0
        status, buyer = status_and_buyer(row[6], row[7] if len(row) > 7 else "")
        laycan_start, laycan_end = month_dates(year, month)
        notes = clean(row[10] if len(row) > 10 else "")
        purchase_date = clean(row[11] if len(row) > 11 else "") or f"{year}-{month:02d}-01"

        cargo = {
            "id": f"x{year}-{row_index:03d}",
            "year": year,
            "month": month,
            "laycanStart": laycan_start,
            "laycanEnd": laycan_end,
            "purchaseType": purchase_type,
            "mine": mine,
            "spec": spec,
            "quantityMt": quantity,
            "status": status,
            "buyer": buyer,
            "purchaseDate": purchase_date,
            "sourceSheet": "2026구매계획 요약",
            "sourceRow": row_index,
        }

        if notes:
            cargo["notes"] = notes
        if purchase_type == "fixed":
            cargo["priceFixed"] = price_fixed or 0
        else:
            cargo["priceFormula"] = price_formula or ""
            cargo["indexRule"] = normalize_rule(row[10] if len(row) > 10 else "")
            previous_rule = clean(row[4])
            if "before" in clean(row[10]).lower():
                cargo["indexRule"] = normalize_rule(row[10])
            elif len(row) > 4:
                cargo["indexRule"] = "1 month before laycan"

        cargos.append(cargo)

    return cargos


def append_latest_ici(indices):
    latest = {
        "id": "ici_2026-05-08",
        "weekStart": "2026-05-04",
        "weekEnd": "2026-05-10",
        "ici1": 122.10,
        "ici2": 94.35,
        "ici3": 80.74,
        "ici4": 63.56,
        "ici5": 37.78,
        "publicationDate": "2026-05-08",
    }
    filtered = [item for item in indices if item.get("publicationDate") != latest["publicationDate"]]
    filtered.append(latest)
    return sorted(filtered, key=lambda item: item["publicationDate"])


def import_ici_from_workbook(wb):
    """Read 'ICI주간데이터' sheet from the uploaded workbook.
    Returns list of ICI index dicts ready to merge into data-store indices.
    Skips rows where column B is not a valid YYYY-MM-DD date (e.g. average rows).
    """
    candidate_names = ["ICI주간데이터", "ICI데이터", "ICI", "ici"]
    ws = None
    for name in candidate_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        return []

    new_entries = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # row 1-3 = title/header
        date_val = row[1] if len(row) > 1 else None        # column B = 발표일
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val).strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
            continue  # skip average rows / headers / blank rows

        def to_float(v):
            if v is None:
                return None
            try:
                return round(float(v), 4)
            except (TypeError, ValueError):
                return None

        # columns: A=월, B=발표일, C=중국내수, D=ICI1, E=ICI2, F=ICI3, G=ICI4, H=ICI5
        ici1 = to_float(row[3] if len(row) > 3 else None)
        ici2 = to_float(row[4] if len(row) > 4 else None)
        ici3 = to_float(row[5] if len(row) > 5 else None)
        ici4 = to_float(row[6] if len(row) > 6 else None)
        ici5 = to_float(row[7] if len(row) > 7 else None)

        # require at least ICI2 or ICI3 to be present
        if ici2 is None and ici3 is None:
            continue

        entry = {
            "id": f"ici_{date_str}",
            "publicationDate": date_str,
        }
        if ici1 is not None:
            entry["ici1"] = ici1
        if ici2 is not None:
            entry["ici2"] = ici2
        if ici3 is not None:
            entry["ici3"] = ici3
        if ici4 is not None:
            entry["ici4"] = ici4
        if ici5 is not None:
            entry["ici5"] = ici5
        new_entries.append(entry)

    return new_entries


def merge_ici_entries(existing, new_entries):
    """Upsert new_entries into existing indices list.
    New dates are appended; existing dates are updated.
    Returns sorted list.
    """
    by_date = {item["publicationDate"]: item for item in existing}
    added, updated = 0, 0
    for entry in new_entries:
        date_str = entry["publicationDate"]
        if date_str in by_date:
            by_date[date_str].update(entry)
            updated += 1
        else:
            by_date[date_str] = entry
            added += 1
    return sorted(by_date.values(), key=lambda x: x["publicationDate"]), added, updated


def prefer_current_indices(reference_indices, current_indices):
    """Keep the live store's ICI rows when importing cargo data.
    The reference file is a seed and can be older than manually updated ICI data.
    """
    by_date = {item["publicationDate"]: item for item in reference_indices}
    for item in current_indices:
        if item.get("publicationDate"):
            by_date[item["publicationDate"]] = item
    return sorted(by_date.values(), key=lambda x: x["publicationDate"])


def formula_key(value):
    return re.sub(r"\s+", "", clean(value).lower().replace("(cfr)", ""))


def apply_business_overrides(imported):
    for cargo in imported:
        if cargo.get("purchaseType") != "index":
            continue
        source_row = cargo.get("sourceRow")
        mine = clean(cargo.get("mine")).lower()
        formula = formula_key(cargo.get("priceFormula"))
        if source_row == 125 and mine == "bayan" and "3800" in clean(cargo.get("spec")):
            cargo.update({
                "year": 2026,
                "month": 8,
                "laycanStart": "2026-08-22",
                "laycanEnd": "2026-08-31",
                "status": "forsale",
                "buyer": "",
                "purchaseDate": "2025-11-12",
                "notes": "8.22-31",
                "priceFormula": "ICI4*0.5+ICI5*0.5+3",
                "indexRule": "1 month before laycan",
            })
        if mine == "mbl" and formula.startswith("ici2prorate+6"):
            cargo["indexRule"] = "3 weeks before laycan"
            parsed_laycan = parse_laycan_note(cargo.get("year", 2026), cargo.get("notes"))
            if parsed_laycan:
                cargo["laycanStart"], cargo["laycanEnd"] = parsed_laycan


def apply_reference_rules(imported, reference_cargos):
    exact = {}
    loose = {}
    for cargo in reference_cargos:
        if cargo.get("year") != 2026 or cargo.get("purchaseType") != "index":
            continue
        rule = cargo.get("indexRule")
        if not rule:
            continue
        exact[(cargo.get("month"), clean(cargo.get("mine")).lower(), clean(cargo.get("spec")).lower(), formula_key(cargo.get("priceFormula")))] = rule
        loose.setdefault((cargo.get("month"), clean(cargo.get("mine")).lower(), clean(cargo.get("spec")).lower()), rule)

    for cargo in imported:
        if cargo.get("purchaseType") != "index":
            continue
        key = (cargo.get("month"), clean(cargo.get("mine")).lower(), clean(cargo.get("spec")).lower(), formula_key(cargo.get("priceFormula")))
        loose_key = (cargo.get("month"), clean(cargo.get("mine")).lower(), clean(cargo.get("spec")).lower())
        cargo["indexRule"] = exact.get(key) or loose.get(loose_key) or cargo.get("indexRule") or "1 month before laycan"


def main():
    base = json.loads(REFERENCE.read_text(encoding="utf-8"))
    existing_store = {}
    try:
        existing_store = json.loads(STORE.read_text(encoding="utf-8"))
    except Exception:
        pass

    # ── cargo import ──────────────────────────────────────────────────────────
    workbook_cargos = import_2026_cargos()
    apply_reference_rules(workbook_cargos, base["cargos"])
    apply_business_overrides(workbook_cargos)
    base["cargos"] = [cargo for cargo in base["cargos"] if cargo.get("year") != 2026] + workbook_cargos

    # ── ICI import (from "ICI주간데이터" sheet if present) ────────────────────
    ici_entries = getattr(import_2026_cargos, "_ici_entries", [])
    ici_added, ici_updated = 0, 0
    base["indices"] = prefer_current_indices(base["indices"], existing_store.get("indices", []))
    if ici_entries:
        base["indices"], ici_added, ici_updated = merge_ici_entries(base["indices"], ici_entries)
    else:
        # No ICI sheet in the uploaded workbook: keep current live ICI history.
        pass

    # ── preserve contractForecast if it exists ────────────────────────────────
    if "contractForecast" not in base:
        if "contractForecast" in existing_store:
            base["contractForecast"] = existing_store["contractForecast"]

    base["generatedAt"] = datetime.now().isoformat(timespec="seconds")
    base["source"] = {
        "excel": str(INPUT),
        "sheet": "구매계획 / ICI주간데이터",
        "mode": "template-import",
    }
    STORE.write_text(json.dumps(base, ensure_ascii=False, indent=2), encoding="utf-8")
    DATA_JS.write_text("window.CPT_DATA = " + json.dumps(base, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")

    may = [c for c in workbook_cargos if c["month"] == 5]
    latest_ici = base["indices"][-1] if base["indices"] else {}
    print(json.dumps({
        "ok": True,
        "imported2026": len(workbook_cargos),
        "iciSheetFound": len(ici_entries) > 0,
        "iciAdded": ici_added,
        "iciUpdated": ici_updated,
        "iciTotalEntries": len(base["indices"]),
        "latestIci": latest_ici,
        "mayCargos": len(may),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
