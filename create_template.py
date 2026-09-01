"""
Generate KCH_구매계획_ICI_Template.xlsx
  Sheet 1 "구매계획"    — cargo import (compatible with import_excel.py)
  Sheet 2 "ICI주간데이터" — weekly ICI update
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent / "KCH_구매계획_ICI_Template.xlsx"

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY     = "06345C"
BLUE     = "0077B6"
MID_BLUE = "1B3A5C"
YELLOW   = "FEF9C3"
ACTUAL   = "F0F9FF"   # light blue  – confirmed Argus data
FORECAST = "FFFBEB"   # light amber – user forecast
AVG_BG   = "1B3A5C"   # dark navy   – monthly average row
WHITE    = "FFFFFF"
SOFT     = "F8FAFC"
GRAY     = "94A3B8"

def thin_border(color="D7E7F0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, text, bg=NAVY, fg=WHITE, sz=10, bold=True, wrap=False):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, size=sz, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def data_cell(cell, value, bg=WHITE, fg="1E293B", bold=False, num_fmt=None, align="center"):
    cell.value = value
    cell.font = Font(name="Arial", size=9, bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if num_fmt:
        cell.number_format = num_fmt


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 1 – 구매계획
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "구매계획"

CARGO_COLS = [
    ("기간",           16, "예: 2026년1월"),
    ("매입유형",        12, "Fixed / Index"),
    ("광산",           14, "예: Adaro"),
    ("규격",           12, "예: NAR4400"),
    ("가격/공식",       22, "고정가 숫자 또는 ICI공식"),
    ("수량(MT)",        12, "예: 80000"),
    ("매도/상태",       16, "Kowepo / Unsold / Planned"),
    ("최종수요자",      14, "최종수요자명"),
    ("(예비A)",         8,  ""),
    ("(예비B)",         8,  ""),
    ("비고/선적창",     18, "예: 6.26-7.5"),
    ("매입일",         14, "예: 2026-01-15"),
]

# Row 1 – title
ws.merge_cells("A1:L1")
ws["A1"].value = "KCH Purchase Plan — 구매계획 입력 시트"
ws["A1"].font  = Font(name="Arial", bold=True, size=13, color=WHITE)
ws["A1"].fill  = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Row 2 – notice
ws.merge_cells("A2:L2")
ws["A2"].value = (
    "※ 3행부터 데이터 입력  |  기간: '2026년1월' 형식  |  "
    "매입유형: Fixed 또는 Index  |  Index면 가격/공식란에 ICI공식 입력"
)
ws["A2"].font  = Font(name="Arial", italic=True, size=9, color="526B82")
ws["A2"].fill  = PatternFill("solid", fgColor=YELLOW)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

# Row 3 – column headers  (import reads min_row=3 → first data row is 4 in template,
#                           but we put headers in row 3 and set min_row=4 in updated script)
for ci, (label, width, _) in enumerate(CARGO_COLS, 1):
    cell = ws.cell(row=3, column=ci)
    hdr(cell, label, bg=BLUE)
    ws.column_dimensions[get_column_letter(ci)].width = width

ws.row_dimensions[3].height = 22

# Row 4 – example hint (light yellow)
hints = [item[2] for item in CARGO_COLS]
for ci, hint in enumerate(hints, 1):
    cell = ws.cell(row=4, column=ci)
    cell.value = hint
    cell.font  = Font(name="Arial", italic=True, size=8, color="94A3B8")
    cell.fill  = PatternFill("solid", fgColor=YELLOW)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border("E2E8F0")
ws.row_dimensions[4].height = 18

# Rows 5-54 – blank data rows
for r in range(5, 55):
    bg = WHITE if r % 2 == 1 else SOFT
    for ci in range(1, 13):
        cell = ws.cell(row=r, column=ci)
        cell.fill   = PatternFill("solid", fgColor=bg)
        cell.border = thin_border("E2E8F0")
        cell.font   = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════════════════════
#  Sheet 2 – ICI주간데이터
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("ICI주간데이터")

ICI_COLS = [
    ("월",          9,  "월 (자동)"),
    ("발표일",      14, "YYYY-MM-DD"),
    ("중국내수\nNAR5500/Yuan", 12, ""),
    ("ICI1",        9,  "$/t"),
    ("ICI2",        9,  "$/t"),
    ("ICI3",        9,  "$/t"),
    ("ICI4",        9,  "$/t"),
    ("ICI5",        9,  "$/t"),
]

# Row 1 – title
ws2.merge_cells("A1:H1")
ws2["A1"].value = "KCH Purchase Plan — ICI 주간 데이터 업데이트 시트"
ws2["A1"].font  = Font(name="Arial", bold=True, size=13, color=WHITE)
ws2["A1"].fill  = PatternFill("solid", fgColor=NAVY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Row 2 – legend
ws2.merge_cells("A2:H2")
ws2["A2"].value = (
    "※ 매주 Argus 발표 후 새 행을 추가하고 발표일(YYYY-MM-DD)+ICI1~5 입력 → 파일 저장 → 시스템 Import  "
    "| 배경 하늘색=실제발표, 노란색=예측"
)
ws2["A2"].font  = Font(name="Arial", italic=True, size=9, color="526B82")
ws2["A2"].fill  = PatternFill("solid", fgColor=YELLOW)
ws2["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[2].height = 18

# Row 3 – column headers
for ci, (label, width, _) in enumerate(ICI_COLS, 1):
    cell = ws2.cell(row=3, column=ci)
    hdr(cell, label, bg=BLUE, wrap=True)
    ws2.column_dimensions[get_column_letter(ci)].width = width
ws2.row_dimensions[3].height = 32

# ── 2026 data ─────────────────────────────────────────────────────────────────
# (date, china_domestic, ici1, ici2, ici3, ici4, ici5, is_forecast)
DATA_2026 = [
    # Jan 2026
    ("2026-01-02", 689,  102.15, 80.15, 60.91, 45.46, 30.97, False),
    ("2026-01-09", 701,  102.35, 80.55, 61.54, 46.23, 31.11, False),
    ("2026-01-16", 704,  102.76, 80.96, 61.84, 46.80, 31.42, False),
    ("2026-01-23", 691,  103.21, 81.39, 61.96, 47.31, 31.74, False),
    ("2026-01-30", 695,  104.12, 81.88, 63.26, 47.70, 32.07, False),
    # Feb 2026
    ("2026-02-06", 696,  104.96, 82.93, 65.80, 49.27, 32.80, False),
    ("2026-02-13", 712,  106.03, 84.04, 67.02, 51.59, 33.76, False),
    ("2026-02-20", 712,  107.30, 85.16, 67.86, 52.64, 34.33, False),
    ("2026-02-27", 742,  108.30, 86.66, 70.35, 55.39, 35.15, False),
    # Mar 2026
    ("2026-03-06", 750,  111.11, 87.71, 72.63, 58.21, 35.84, False),
    ("2026-03-13", 735,  114.02, 88.92, 73.62, 59.97, 36.03, False),
    ("2026-03-20", 757,  115.36, 88.96, 74.21, 60.03, 36.13, False),
    ("2026-03-27", 756,  116.94, 89.75, 74.89, 60.31, 36.38, False),
    # Apr 2026
    ("2026-04-03", 759,  117.78, 90.34, 75.47, 60.51, 36.24, False),
    ("2026-04-10", 763,  118.70, 91.00, 76.50, 60.40, 36.48, False),
    ("2026-04-17", 769,  118.86, 91.55, 77.43, 60.69, 36.44, False),
    ("2026-04-24", 773,  119.44, 91.67, 77.73, 60.85, 36.89, False),
    # May 2026 – actual
    ("2026-05-01", 804,  120.99, 92.88, 79.07, 61.82, 37.46, False),
    ("2026-05-08", 810,  122.10, 94.35, 80.74, 63.56, 37.78, False),
    # May 2026 – forecast
    ("2026-05-15", 820,  122.98, 102.21, 82.74, 64.32, 39.48, True),
    ("2026-05-22", 830,  123.58, 103.21, 83.74, 64.82, 39.98, True),
    ("2026-05-29", 850,  124.18, 104.21, 84.74, 65.32, 40.48, True),
    # Jun 2026 – forecast
    ("2026-06-05", 850,  124.78, 104.91, 85.74, 65.82, 40.98, True),
    ("2026-06-12", 850,  125.38, 105.61, 86.74, 66.32, 41.48, True),
    ("2026-06-19", 850,  125.98, 106.31, 87.74, 66.82, 41.98, True),
    ("2026-06-26", 850,  126.58, 107.01, 88.74, 67.32, 42.48, True),
    # Jul 2026 – forecast
    ("2026-07-03", 804,  127.18, 107.71, 89.14, 67.82, 42.68, True),
    ("2026-07-10", 810,  127.78, 108.41, 89.54, 68.32, 42.88, True),
    ("2026-07-17", 820,  128.38, 109.11, 89.94, 68.82, 43.08, True),
    ("2026-07-24", 830,  128.98, 109.81, 90.34, 69.32, 43.28, True),
    ("2026-07-31", 850,  129.58, 110.51, 90.74, 69.82, 43.48, True),
    # Aug 2026 – forecast
    ("2026-08-07", 850,  128.79, 103.18, 85.47, 70.32, 40.06, True),
    ("2026-08-14", 850,  128.79, 103.18, 85.47, 70.82, 40.06, True),
    ("2026-08-21", 850,  128.79, 103.18, 85.47, 71.32, 40.06, True),
    ("2026-08-28", 850,  128.79, 103.18, 85.47, 71.82, 40.06, True),
]

MONTH_LABELS = {
    "01": "Jan 2026", "02": "Feb 2026", "03": "Mar 2026", "04": "Apr 2026",
    "05": "May 2026", "06": "Jun 2026", "07": "Jul 2026", "08": "Aug 2026",
}

# group by month to build avg rows
from itertools import groupby

def month_key(row): return row[0][5:7]

current_row = 4
month_groups = {}  # month_str -> (first_data_row, last_data_row)

prev_month = None
month_first_row = None

for row_data in DATA_2026:
    date_str, china_dom, ici1, ici2, ici3, ici4, ici5, is_fc = row_data
    mo = date_str[5:7]

    # insert average row when month changes
    if prev_month and mo != prev_month:
        # write average row
        avg_first = month_groups[prev_month][0]
        avg_last  = current_row - 1
        r = current_row
        label_col = get_column_letter(1)
        ws2.cell(row=r, column=1).value = f"{MONTH_LABELS.get(prev_month, prev_month)} 평균"
        ws2.cell(row=r, column=1).font  = Font(name="Arial", bold=True, size=9, color=WHITE)
        ws2.cell(row=r, column=1).fill  = PatternFill("solid", fgColor=MID_BLUE)
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row=r, column=1).border = thin_border("FFFFFF")

        ws2.cell(row=r, column=2).value = ""
        ws2.cell(row=r, column=2).fill  = PatternFill("solid", fgColor=MID_BLUE)
        ws2.cell(row=r, column=2).border = thin_border("FFFFFF")

        for ci, col_letter in enumerate([None, None, "C", "D", "E", "F", "G", "H"], 1):
            if ci <= 2: continue
            cell = ws2.cell(row=r, column=ci)
            cell.value = f"=AVERAGE({col_letter}{avg_first}:{col_letter}{avg_last})"
            cell.font  = Font(name="Arial", bold=True, size=9, color=WHITE)
            cell.fill  = PatternFill("solid", fgColor=MID_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border("FFFFFF")
            cell.number_format = "0.00"

        ws2.row_dimensions[r].height = 18
        current_row += 1

    if mo != prev_month:
        month_groups[mo] = [current_row, current_row]
        prev_month = mo
    else:
        month_groups[mo][1] = current_row

    # write data row
    bg = FORECAST if is_fc else ACTUAL
    r = current_row

    # Col A – month label (only on first row of month)
    is_first = (month_groups[mo][0] == r)
    ws2.cell(row=r, column=1).value = MONTH_LABELS.get(mo, mo) if is_first else ""
    ws2.cell(row=r, column=1).font  = Font(name="Arial", bold=is_first, size=9,
                                            color=NAVY if not is_fc else "92400E")
    ws2.cell(row=r, column=1).fill  = PatternFill("solid", fgColor=bg)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=r, column=1).border = thin_border()

    # Col B – publication date
    data_cell(ws2.cell(row=r, column=2), date_str, bg=bg, fg="0F172A", bold=is_fc)

    # Col C – China Domestic
    data_cell(ws2.cell(row=r, column=3), china_dom, bg=bg, num_fmt="#,##0")

    # Cols D-H – ICI1-5
    for ci, val in enumerate([ici1, ici2, ici3, ici4, ici5], 4):
        c = ws2.cell(row=r, column=ci)
        data_cell(c, val, bg=bg,
                  fg="0F172A" if not is_fc else "92400E",
                  bold=is_fc, num_fmt="0.00")

    ws2.row_dimensions[r].height = 17
    current_row += 1

# final month average row
if prev_month:
    avg_first = month_groups[prev_month][0]
    avg_last  = current_row - 1
    r = current_row
    ws2.cell(row=r, column=1).value = f"{MONTH_LABELS.get(prev_month, prev_month)} 평균"
    ws2.cell(row=r, column=1).font  = Font(name="Arial", bold=True, size=9, color=WHITE)
    ws2.cell(row=r, column=1).fill  = PatternFill("solid", fgColor=MID_BLUE)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=r, column=1).border = thin_border("FFFFFF")
    ws2.cell(row=r, column=2).fill  = PatternFill("solid", fgColor=MID_BLUE)
    ws2.cell(row=r, column=2).border = thin_border("FFFFFF")
    for ci, col_letter in enumerate([None, None, "C", "D", "E", "F", "G", "H"], 1):
        if ci <= 2: continue
        cell = ws2.cell(row=r, column=ci)
        cell.value = f"=AVERAGE({col_letter}{avg_first}:{col_letter}{avg_last})"
        cell.font  = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border("FFFFFF")
        cell.number_format = "0.00"
    ws2.row_dimensions[r].height = 18
    current_row += 1

# ── "새 데이터 추가" blank rows ────────────────────────────────────────────
ws2.merge_cells(f"A{current_row}:H{current_row}")
ws2.cell(row=current_row, column=1).value = "▼ 새 주간 데이터를 아래에 추가하세요 (발표일+ICI1~5 입력 후 저장)"
ws2.cell(row=current_row, column=1).font  = Font(name="Arial", italic=True, size=9, color="64748B")
ws2.cell(row=current_row, column=1).fill  = PatternFill("solid", fgColor="F1F5F9")
ws2.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[current_row].height = 20
current_row += 1

for r in range(current_row, current_row + 12):
    bg = WHITE if r % 2 == 0 else SOFT
    for ci in range(1, 9):
        cell = ws2.cell(row=r, column=ci)
        cell.fill   = PatternFill("solid", fgColor=bg)
        cell.border = thin_border("E2E8F0")
        cell.font   = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ci in (4, 5, 6, 7, 8):
            cell.number_format = "0.00"
    ws2.row_dimensions[r].height = 17

ws2.freeze_panes = "A4"

# ── Legend box (right side, column J) ─────────────────────────────────────
legend = [
    ("범례 (Legend)", NAVY, WHITE, True),
    ("하늘색 = Argus 실제 발표 (수정 금지)", ACTUAL, NAVY, False),
    ("노란색 = 예측값 (수정 가능)", FORECAST, "92400E", False),
    ("짙은 파랑 = 월평균 (자동계산)", MID_BLUE, WHITE, False),
]
ws2.column_dimensions["J"].width = 32
for i, (text, bg, fg, bold) in enumerate(legend, 3):
    cell = ws2.cell(row=i, column=10)
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, size=9, color=fg)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border()
    ws2.row_dimensions[i].height = 18

wb.save(OUT)
print(f"Template saved → {OUT}")
